from minio.error import S3Error
from minio import Minio
import streamlit as st
import io
from PyPDF2 import PdfReader, PdfWriter
import os
from config.settings import MINIO_URL, MINIO_BUCKET_NAME, MINIO_USERNAME, MINIO_PASSWORD, TEMP_DIR
import shutil
import xlwings as xw
import time
import subprocess
from openpyxl import load_workbook, Workbook

def save_uploaded_file(uploaded_file):
    """Save uploaded file to temporary directory."""
    os.makedirs(TEMP_DIR, exist_ok=True)
    file_path = os.path.join(TEMP_DIR, uploaded_file.name)
    with open(file_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return file_path

def cleanup_file(file_path):
    """Remove file or directory from temporary directory."""
    if os.path.isdir(file_path):
        shutil.rmtree(file_path)
    elif os.path.isfile(file_path):
        os.remove(file_path)

# ------------MINIO
def connect_to_minio():
    client = Minio(MINIO_URL,
                access_key=MINIO_USERNAME,
                secret_key=MINIO_PASSWORD,
                secure=False)

    # Create the bucket (create if not exists)
    try:
        if not client.bucket_exists(MINIO_BUCKET_NAME):
            print("Bucket does not exist. Creating...")
            try:
                client.make_bucket(MINIO_BUCKET_NAME)
            except Exception as e:
                pass
                # st.error(f"Failed to create main bucket: {e}")
        else:
            print("Bucket already exists.")
            pass    
    except Exception as e:
        pass
        # st.error(f"Error checking main bucket: {e}")

    return client

def create_folder(client, bucket, folder_name):
    try:
        # Tạo object rỗng với tên "folder/"
        object_name = folder_name.rstrip("/") + "/placeholder.txt"
        client.put_object(
            bucket,
            object_name,
            data=io.BytesIO(b""),  # empty
            length=0,
            content_type="text/plain"
        )
        print(f"✅ Created folder: {folder_name}")
    except S3Error as e:
        print(f"❌ Failed to create folder '{folder_name}': {e}")

def list_folders(client, bucket):
    folders = set()
    try:
        objects = client.list_objects(bucket, recursive=False)
        for obj in objects:
            if "/" in obj.object_name:
                folder = obj.object_name.split("/")[0]
                folders.add(folder)
        return sorted(list(folders))
    except S3Error as e:
        # st.error(f"Error listing folders: {e}")
        return []


def list_level_2_folders(client, bucket, folder_name):
    folders = set()
    try:
        if not folder_name.endswith("/"):
            folder_name += "/"

        objects = client.list_objects(bucket, prefix=folder_name, recursive=True)
        for obj in objects:
            parts = obj.object_name.split("/")
            if len(parts) >= 2:
                level2_folder = "/".join(parts[:2])  
                folders.add(level2_folder)
        return sorted(folders)
    except S3Error as e:
        print(f"❌ MinIO error: {e}")
        return []

def remove_objects(client, bucket, userid, subject=None, level=None):
    try:
        objects = client.list_objects(bucket, prefix=f"{userid}/", recursive=True)

        for obj in objects:
            object_name = obj.object_name
            if subject and level:
                if subject in object_name and level in object_name:
                    client.remove_object(bucket, object_name)
            else:
                client.remove_object(bucket, object_name)

        return True
    except S3Error as err:
        return False

def upload_file(client, bucket, file_path, object_name):
    try:
        client.fput_object(bucket, object_name, file_path)
        return True
    except S3Error as err:
        return False

def split_pdf_by_titles(input_pdf_path, document_titles, client, bucket_name, output_dir="split_docs"):

    os.makedirs(output_dir, exist_ok=True)
    reader = PdfReader(input_pdf_path)
    total_pages = len(reader.pages)
    uploaded_files = []

    for doc in document_titles:
        title = doc["title"]
        folder = doc["folder_recommendation"]
        pages = doc["page_numbers/ sheet numbers"]

        writer = PdfWriter()
        for p in pages:
            if 1 <= p <= total_pages:
                writer.add_page(reader.pages[p - 1])
            else:
                print(f"[WARNING] Page {p} out of range for title: {title}")

        # Tạo tên file an toàn
        safe_title = title.replace(" ", "_").replace("/", "_").replace("\\", "_")[:100]
        filename = f"{safe_title}.pdf"
        local_folder_path = os.path.join(output_dir, folder)
        os.makedirs(local_folder_path, exist_ok=True)

        local_file_path = os.path.join(local_folder_path, filename)
        with open(local_file_path, "wb") as f:
            writer.write(f)

        # Upload lên MinIO
        object_name = f"{folder}/{filename}"
        try:
            client.fput_object(bucket_name, object_name, local_file_path)
            uploaded_files.append(object_name)
        except S3Error as e:
            print(f"❌ Error uploading {filename}: {e}")

    # Clean up 
    cleanup_file(output_dir)
    return uploaded_files

# def split_xlsx(excel_file):
#     app = xw.App(visible=False)
#     wb = app.books.open(excel_file)
#     for sheet in wb.sheets:
#         try: 
#             sheet.api.Copy()
#             wb_new = xw.books.active
#             wb_new.save(f"{sheet.name}.xlsx")
#             wb_new.close()
#         except Exception as e:
#             print(f"Error when processing file: {sheet.name}")
#     wb.close()
#     app.quit()
# def split_xlsx_by_titles(input_xlsx_path, document_titles, client, bucket_name, output_dir="split_docs"):

#     os.makedirs(output_dir, exist_ok=True)
#     app = xw.App(visible=False)
#     try:
#         wb = app.books.open(input_xlsx_path)
#         uploaded_files = []
        
#         title_map = {}
#         for doc in document_titles:
#             title = doc["title"]
#             sheets = doc["page_numbers/ sheet numbers"]
#             folder = doc["folder_recommendation"]
#             for sheet_ref in sheets:
#                 title_map[sheet_ref] = (folder,title)

     
#         for sheet_idx, sheet in enumerate(wb.sheets, 1):
#             try:
#                 folder, safe_title = title_map[sheet_idx]
#                 safe_title = safe_title.replace(" ", "_").replace("/", "_").replace("\\", "_")[:100]
#                 filename = f"{safe_title}.xlsx"
                
#                 sheet.api.Copy()
#                 wb_new = xw.books.active
                
#                 local_folder_path = os.path.join(output_dir, folder)
#                 os.makedirs(local_folder_path, exist_ok=True)
#                 local_file_path = os.path.join(local_folder_path, filename)
#                 wb_new.save(local_file_path)
#                 wb_new.close()
#                 # Upload to MinIO
#                 object_name = f"{folder}/{filename}"
#                 try:
#                     client.fput_object(bucket_name, object_name, local_file_path)
#                     uploaded_files.append(object_name)
#                 except S3Error as e:
#                     print(f"❌ Error uploading {filename}: {e}")
#             except Exception as e:
#                 print(f"[WARNING] Error processing sheet {sheet.name}: {e}")
        
#         wb.close()
#     finally:
#         app.quit()

#     cleanup_file(output_dir)
#     return uploaded_files

def convert_xlsx_to_pdf(input_path, output_dir):
    result = subprocess.run([
        "libreoffice", "--headless", "--convert-to", "pdf", input_path, "--outdir", output_dir
    ], capture_output=True, text=True)
    
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice convert failed: {result.stderr}")
    
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    return os.path.join(output_dir, f"{base_name}.pdf")

def split_xlsx_by_titles(input_xlsx_path, document_titles, client, bucket_name, output_dir="split_docs"):
    os.makedirs(output_dir, exist_ok=True)
    uploaded_files = []

    wb = load_workbook(filename=input_xlsx_path)
    
    # Map: sheet_index → (folder, title)
    title_map = {}
    for doc in document_titles:
        title = doc["title"]
        sheets = doc["page_numbers/ sheet numbers"]
        folder = doc["folder_recommendation"]
        for sheet_ref in sheets:
            title_map[sheet_ref] = (folder, title)

    for sheet_idx, sheetname in enumerate(wb.sheetnames, 1):
        try:
            if sheet_idx not in title_map:
                print(f"⚠️ Skip sheet {sheetname} – no mapping")
                continue

            folder, title = title_map[sheet_idx]
            safe_title = title.replace(" ", "_").replace("/", "_").replace("\\", "_")[:100]
            local_folder_path = os.path.join(output_dir, folder)
            os.makedirs(local_folder_path, exist_ok=True)
            temp_xlsx_path = os.path.join(local_folder_path, f"{safe_title}.xlsx")

            # Create new Excel file with 1 sheet
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = sheetname

            source_ws = wb[sheetname]
            for row in source_ws.iter_rows(values_only=True):
                new_ws.append(row)

            new_wb.save(temp_xlsx_path)
            new_wb.close()

            # Convert to PDF
            pdf_path = convert_xlsx_to_pdf(temp_xlsx_path, local_folder_path)
            object_name = f"{folder}/{safe_title}.pdf"

            # Upload PDF
            try:
                client.fput_object(bucket_name, object_name, pdf_path)
                uploaded_files.append(object_name)
            except S3Error as e:
                print(f"❌ Error uploading {object_name}: {e}")
        except Exception as e:
            print(f"[WARNING] Error processing sheet {sheetname}: {e}")

    wb.close()
    cleanup_file(output_dir)
    return uploaded_files


if __name__ == "__main__":
    client = connect_to_minio()
    create_folder(client, MINIO_BUCKET_NAME, "Material_Acceptance_Forms")
    create_folder(client, MINIO_BUCKET_NAME, "Material_Quantity_Lists")
    print(list_folders(client, MINIO_BUCKET_NAME))
    print(list_level_2_folders(client, MINIO_BUCKET_NAME, "Material_Acceptance_Forms"))
    